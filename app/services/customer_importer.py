from __future__ import annotations

import csv
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from fastapi import UploadFile

from app.core.config import settings
from app.core import db


SUPPORTED_EXTENSIONS = {".csv", ".tsv", ".xlsx", ".xlsm"}
PHONE_KEYWORDS = ("phone", "mobile", "tel", "number", "手机号", "电话", "手机", "号码", "微信号")
GREETING_KEYWORDS = ("greeting", "verify", "招呼", "验证", "申请", "备注语")
REMARK_KEYWORDS = ("remark", "name", "customer", "备注", "姓名", "名称", "客户", "昵称")


@dataclass
class ParsedSheet:
    headers: list[dict[str, Any]]
    rows: list[dict[str, Any]]
    row_count: int
    header_row: int


def save_upload(file: UploadFile) -> dict[str, str]:
    settings.ensure_dirs()
    folder = settings.uploads_dir / "customers"
    folder.mkdir(parents=True, exist_ok=True)
    safe_name = Path(file.filename or "customers.xlsx").name
    suffix = Path(safe_name).suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        raise ValueError("Only .xlsx, .xlsm, .csv, and .tsv files are supported.")
    target = folder / f"{uuid.uuid4().hex}_{safe_name}"
    with target.open("wb") as output:
        while chunk := file.file.read(1024 * 1024):
            output.write(chunk)
    return {"path": str(target), "filename": safe_name}


def preview_file(path: str, sheet_name: str | None = None, limit: int = 20) -> dict[str, Any]:
    source = Path(path)
    parsed, sheets = _parse_file(source, sheet_name=sheet_name, limit=limit)
    return {
        "path": str(source),
        "filename": source.name.split("_", 1)[-1],
        "sheets": sheets,
        "sheet_name": sheet_name or (sheets[0] if sheets else None),
        "headers": parsed.headers,
        "rows": parsed.rows,
        "row_count": parsed.row_count,
        "header_row": parsed.header_row,
        "suggested_mapping": suggest_mapping(parsed.headers),
    }


def import_file(path: str, sheet_name: str | None, mapping: dict[str, str | None]) -> dict[str, Any]:
    source = Path(path)
    parsed, sheets = _parse_file(source, sheet_name=sheet_name, limit=None)
    effective_sheet = sheet_name or (sheets[0] if sheets else None)
    import_id = uuid.uuid4().hex
    records = []
    for row in parsed.rows:
        raw = row.get("raw") or {}
        number = raw.get(mapping.get("number") or "")
        if number is None:
            number = row.get(mapping.get("number") or "")
        records.append(
            {
                "number": number,
                "greetings": raw.get(mapping.get("greetings") or "") or "",
                "remark": raw.get(mapping.get("remark") or "") or "",
                "raw": raw,
                "source_filename": source.name.split("_", 1)[-1],
                "source_sheet": effective_sheet,
                "source_row": row.get("_row_number"),
                "import_id": import_id,
            }
        )
    counts = db.upsert_customers(records)
    import_row = db.create_customer_import(
        import_id=import_id,
        filename=source.name.split("_", 1)[-1],
        path=str(source),
        sheet_name=effective_sheet,
        mapping=mapping,
        row_count=parsed.row_count,
        imported_count=counts["imported_count"],
        updated_count=counts["updated_count"],
        skipped_count=counts["skipped_count"],
    )
    return {**import_row, **counts}


def suggest_mapping(headers: list[dict[str, Any]]) -> dict[str, str | None]:
    labels = [(header["key"], str(header["label"]).lower()) for header in headers]
    return {
        "number": _first_match(labels, PHONE_KEYWORDS),
        "greetings": _first_match(labels, GREETING_KEYWORDS),
        "remark": _first_match(labels, REMARK_KEYWORDS),
    }


def _first_match(labels: list[tuple[str, str]], keywords: tuple[str, ...]) -> str | None:
    for keyword in keywords:
        lowered = keyword.lower()
        for key, label in labels:
            if lowered in label:
                return key
    return None


def _parse_file(source: Path, sheet_name: str | None, limit: int | None) -> tuple[ParsedSheet, list[str]]:
    suffix = source.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        return _parse_csv(source, limit), [source.stem]
    if suffix in {".xlsx", ".xlsm"}:
        return _parse_xlsx(source, sheet_name, limit)
    raise ValueError("Only .xlsx, .xlsm, .csv, and .tsv files are supported.")


def _parse_csv(source: Path, limit: int | None) -> ParsedSheet:
    delimiter = "\t" if source.suffix.lower() == ".tsv" else ","
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            with source.open("r", encoding=encoding, newline="") as handle:
                rows = list(csv.reader(handle, delimiter=delimiter))
            break
        except UnicodeDecodeError:
            rows = []
    else:
        rows = []
    return _rows_to_sheet(rows, limit)


def _parse_xlsx(source: Path, sheet_name: str | None, limit: int | None) -> tuple[ParsedSheet, list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to parse Excel files. Run `pip install -r requirements.txt`.") from exc

    workbook = load_workbook(source, read_only=True, data_only=True)
    sheets = workbook.sheetnames
    if not sheets:
        raise ValueError("Workbook has no sheets.")
    selected = sheet_name if sheet_name in sheets else sheets[0]
    worksheet = workbook[selected]
    rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    workbook.close()
    return _rows_to_sheet(rows, limit), sheets


def _rows_to_sheet(rows: list[list[Any]], limit: int | None) -> ParsedSheet:
    header_index = _find_header_row(rows)
    if header_index is None:
        return ParsedSheet(headers=[], rows=[], row_count=0, header_row=0)
    headers = _build_headers(rows[header_index])
    data_rows = []
    for offset, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        raw = _row_to_raw(headers, row)
        if not any(_clean_value(value) not in ("", None) for value in raw.values()):
            continue
        data_rows.append({"_row_number": offset, "raw": raw, **raw})
        if limit is not None and len(data_rows) >= limit:
            break
    row_count = 0
    for row in rows[header_index + 1 :]:
        raw = _row_to_raw(headers, row)
        if any(_clean_value(value) not in ("", None) for value in raw.values()):
            row_count += 1
    return ParsedSheet(headers=headers, rows=data_rows, row_count=row_count, header_row=header_index + 1)


def _find_header_row(rows: list[list[Any]]) -> int | None:
    for index, row in enumerate(rows[:20]):
        values = [_clean_value(value) for value in row]
        filled = [value for value in values if value not in ("", None)]
        if len(filled) >= 2 or any(any(keyword in str(value) for keyword in PHONE_KEYWORDS) for value in filled):
            return index
    return None


def _build_headers(row: list[Any]) -> list[dict[str, Any]]:
    headers = []
    seen: dict[str, int] = {}
    for index, value in enumerate(row):
        label = str(_clean_value(value) or f"Column {index + 1}").strip()
        count = seen.get(label, 0)
        seen[label] = count + 1
        key = label if count == 0 else f"{label}_{count + 1}"
        headers.append({"key": key, "label": label, "index": index})
    return headers


def _row_to_raw(headers: list[dict[str, Any]], row: list[Any]) -> dict[str, Any]:
    raw = {}
    for header in headers:
        index = int(header["index"])
        raw[header["key"]] = _clean_value(row[index] if index < len(row) else None)
    return raw


def _clean_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return value
